"""
Refresh the s8e8 Google PPC finance snapshot used by build_google_ads_dashboard.py.

The s8e8 "Marketing ROI Report" xlsx is a manually-compiled finance report
(production/collections figures are hand-entered from the practice management
system each month) -- there is no live API for it. Whenever Finance shares a
new export, re-run this script against it to push the parsed data into the
"s8e8_finance_raw" tab of the same Google Sheet google_ads_raw/deals_raw live in
(requires GOOGLE_SA_JSON + GOOGLE_SHEET_ID env vars, same as the other sync
scripts). Raw data is never committed to this repo -- see .gitignore.

    export GOOGLE_SA_JSON='...'
    export GOOGLE_SHEET_ID='...'
    python3 scripts/update_s8e8_snapshot.py "/path/to/s8e8 - Marketing ROI Report.xlsx"

Pass --out-csv to also dump a local CSV copy for inspection (gitignored, not
committed) instead of/alongside writing to the Sheet.
"""
import argparse
import openpyxl, os, re, datetime, csv, json
from pathlib import Path

DEFAULT_LOG = Path(__file__).resolve().parent / "data" / "s8e8_google_ppc_snapshot.log.json"

MONTHS = {m.lower(): i + 1 for i, m in enumerate(
    ['January', 'February', 'March', 'April', 'May', 'June', 'July',
     'August', 'September', 'October', 'November', 'December'])}
MONTH_NAMES = {v: k.capitalize() for k, v in MONTHS.items()}

MARKET_MAP = {
    'Dallas': 'Dallas',
    'Houston': 'Houston',
    'PHX ': 'Phoenix',
    'Tucson': 'Tucson',
    'Austin': 'Austin',
    'San Antonio': 'San Antonio',
    'Utah': 'Utah',
    'PROVO': 'Provo',
}


def is_monthish(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return True
    if isinstance(v, str):
        s = v.strip().lower()
        if s in MONTHS:
            return True
        m = re.match(r'^([a-z]+)\s+(\d{4})$', s)
        if m and m.group(1) in MONTHS:
            return True
    return False


def parse_month_cell(v):
    """Return (month_num, year_or_None, tag) or None if not a month-like cell (skip: YTD etc.)"""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.month, v.year, 'explicit-date')
    if isinstance(v, str):
        s = v.strip()
        low = s.lower()
        m = re.match(r'^([a-z]+)\s+(\d{4})$', low)
        if m and m.group(1) in MONTHS:
            return (MONTHS[m.group(1)], int(m.group(2)), 'explicit-string')
        if low in MONTHS:
            return (MONTHS[low], None, 'bare-month')
    return None


def find_month_row(ws, max_scan_row=13):
    best_row, best_count = None, -1
    for r in range(1, max_scan_row):
        cnt = sum(1 for c in range(1, ws.max_column + 1) if is_monthish(ws.cell(row=r, column=c).value))
        if cnt > best_count:
            best_count, best_row = cnt, r
    return best_row, best_count


def find_google_ppc_row(ws, max_scan_row=80):
    hits = []
    for r in range(1, max_scan_row):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str):
            low = v.lower()
            if 'google' in low and ('ppc' in low or 'adwords' in low):
                hits.append((r, v))
    return hits


def year_marker_scan(ws, month_row, max_col):
    """Look in rows just above month_row for standalone year-like numbers (1990-2035)."""
    markers = []
    for r in range(max(1, month_row - 4), month_row):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                if 1990 <= v <= 2035 and float(v).is_integer():
                    markers.append((c, int(v)))
    markers.sort()
    return markers


def nearest_marker(markers, col):
    best = None
    for c, y in markers:
        if c <= col:
            best = y
        else:
            break
    return best


def combined_header_text(ws, month_row, col, window=4):
    parts = []
    for r in range(month_row + 1, month_row + 1 + window):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
    return ' '.join(parts).lower()


def process_sheet(wb, sheet_name, market_name, log):
    ws = wb[sheet_name]
    max_col = ws.max_column
    month_row, month_count = find_month_row(ws)
    log['month_row'] = month_row
    log['month_row_hit_count'] = month_count

    # block-start columns = columns in month_row that are month-like OR clearly a section label (e.g. "Year to date")
    block_starts = []  # (col, raw_value)
    for c in range(1, max_col + 1):
        v = ws.cell(row=month_row, column=c).value
        if v is not None and (is_monthish(v) or (isinstance(v, str) and v.strip() != '')):
            block_starts.append((c, v))

    if not block_starts:
        log['error'] = 'no block-start columns found'
        return [], log

    # compute block widths (gap to next block-start), last block width = min(gap-guess, 12) capped
    widths = []
    for i, (c, v) in enumerate(block_starts):
        if i + 1 < len(block_starts):
            w = block_starts[i + 1][0] - c
        else:
            w = 12  # generous cap for trailing block
        widths.append(w)

    year_markers = year_marker_scan(ws, month_row, max_col)
    log['year_markers'] = year_markers

    # parse each block-start's month/year
    parsed = []
    for (c, v), w in zip(block_starts, widths):
        pm = parse_month_cell(v)
        skip_reason = None
        if pm is None:
            skip_reason = f'not a month label: {v!r}'
            parsed.append({'col': c, 'width': w, 'raw': v, 'month': None, 'year': None,
                            'tag': None, 'skip_reason': skip_reason})
            continue
        month_num, year, tag = pm
        if year is None:
            ym = nearest_marker(year_markers, c)
            if ym is not None:
                year = ym
                tag = 'year-marker'
        parsed.append({'col': c, 'width': w, 'raw': v, 'month': month_num, 'year': year,
                        'tag': tag, 'skip_reason': None})

    # forward/backward fill years for bare months lacking a year
    n = len(parsed)
    # forward pass
    last_known = None
    for i in range(n):
        p = parsed[i]
        if p['month'] is None:
            continue
        if p['year'] is not None:
            last_known = (p['month'], p['year'])
            continue
        if last_known is not None:
            lm, ly = last_known
            if p['month'] > lm:
                inferred_year = ly
            else:
                inferred_year = ly + 1
            p['year'] = inferred_year
            p['tag'] = 'inferred-forward'
            last_known = (p['month'], p['year'])
    # backward pass for any still missing (leading bare months before first known year)
    last_known = None
    for i in range(n - 1, -1, -1):
        p = parsed[i]
        if p['month'] is None:
            continue
        if p['year'] is not None:
            last_known = (p['month'], p['year'])
            continue
        if last_known is not None:
            lm, ly = last_known
            if p['month'] > lm:
                inferred_year = ly - 1
            else:
                inferred_year = ly
            p['year'] = inferred_year
            p['tag'] = 'inferred-backward'
            last_known = (p['month'], p['year'])

    log['blocks_total'] = len(parsed)
    log['blocks_skipped_non_month'] = sum(1 for p in parsed if p['skip_reason'])
    log['blocks_year_unresolved'] = sum(1 for p in parsed if p['month'] is not None and p['year'] is None)

    # find google ppc row
    hits = find_google_ppc_row(ws)
    log['google_ppc_row_candidates'] = hits
    if not hits:
        log['error'] = 'no Google PPC row found'
        return [], log
    # prefer the one containing both 'ppc' and 'adwords' if multiple, else first
    chosen = None
    for r, label in hits:
        low = label.lower()
        if 'ppc' in low and 'adwords' in low:
            chosen = (r, label)
            break
    if chosen is None:
        chosen = hits[0]
    g_row, g_label = chosen
    log['chosen_google_ppc_row'] = g_row
    log['chosen_google_ppc_label'] = g_label

    is_provo = (market_name == 'Provo')

    rows_out = []
    for p in parsed:
        if p['skip_reason'] or p['month'] is None:
            continue
        c0 = p['col']
        w = p['width']
        cols = list(range(c0, c0 + w))
        # determine role columns via header text, with positional fallback
        patients_col = None
        ad_spend_col = None
        collections_col = None
        cost_per_col = None

        header_texts = {c: combined_header_text(ws, month_row, c) for c in cols}

        if is_provo:
            # Provo-specific: prefer provo-labeled patient & ad-spend columns
            for c in cols:
                t = header_texts[c]
                if 'patient' in t and ('2pro' in t.replace(' ', '') or '#2pro' in t):
                    patients_col = c
                    break
            if patients_col is None:
                for c in cols:
                    if 'patient' in header_texts[c]:
                        patients_col = c
                        break
            for c in cols:
                t = header_texts[c]
                if 'ad spend' in t and 'provo' in t.replace('#', ''):
                    ad_spend_col = c
                    break
            # collections: none exist for Provo tab
            collections_col = None
            cost_per_col = None
        else:
            for c in cols:
                t = header_texts[c]
                if patients_col is None and 'patient' in t:
                    patients_col = c
                elif ad_spend_col is None and ('ad spend' in t or 'spend' in t) and 'total' not in t:
                    ad_spend_col = c
                elif collections_col is None and 'collect' in t:
                    collections_col = c
                elif cost_per_col is None and ('cost per' in t or 'cost/pt' in t or t.strip() == 'cost'):
                    cost_per_col = c

            # positional fallback when no headers found at all in this block (e.g. Utah pre-2016 single columns)
            if patients_col is None and ad_spend_col is None and collections_col is None and cost_per_col is None:
                if w == 1:
                    patients_col = c0
                elif w == 2:
                    patients_col, ad_spend_col = c0, c0 + 1
                elif w == 3:
                    patients_col, ad_spend_col, cost_per_col = c0, c0 + 1, c0 + 2
                elif w >= 4:
                    patients_col, ad_spend_col, collections_col, cost_per_col = c0, c0 + 1, c0 + 2, c0 + 3

        def cellval(col):
            if col is None:
                return None
            v = ws.cell(row=g_row, column=col).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return v
            return None

        patients = cellval(patients_col)
        ad_spend = cellval(ad_spend_col)
        collections = cellval(collections_col)
        cost_per_direct = cellval(cost_per_col)

        notes = []
        if p['tag'] and p['tag'] != 'explicit-date':
            notes.append(f"year_tag={p['tag']}")
        if patients_col is None and ad_spend_col is None:
            notes.append('no data columns identified for this block')

        cost_per_patient = None
        if ad_spend is not None and patients not in (None, 0):
            cost_per_patient = ad_spend / patients
            if cost_per_direct is not None and cost_per_direct != 0:
                pct_diff = abs(cost_per_patient - cost_per_direct) / abs(cost_per_direct)
                if pct_diff > 0.01:
                    notes.append(f'MISMATCH direct_cost_per={cost_per_direct:.4f} computed={cost_per_patient:.4f} pct_diff={pct_diff:.2%}')
        elif cost_per_direct is not None:
            cost_per_patient = cost_per_direct

        if patients == 0 and ad_spend not in (None, 0):
            notes.append('anomaly: patients=0 but spend>0')
        if patients is None and ad_spend is None and collections is None:
            # fully blank block (future/template month with no real inputs yet) -- skip.
            # Note: a stray cost_per_direct==0 can appear here from a formula dividing
            # blank cells (e.g. IFERROR(spend/patients,0)); it is not real data, so it
            # does not count as "has data" for this check.
            continue

        rows_out.append({
            'market': market_name,
            'year': p['year'],
            'month': MONTH_NAMES[p['month']],
            'month_num': p['month'],
            'patients': patients,
            'ad_spend': ad_spend,
            'collections': collections,
            'cost_per_patient': cost_per_patient,
            'source_row': g_row,
            'block_width': w,
            'notes': '; '.join(notes) if notes else '',
        })

    return rows_out, log


def write_to_sheet(all_rows):
    import gspread
    from google.oauth2.service_account import Credentials

    creds = Credentials.from_service_account_info(
        json.loads(os.environ["GOOGLE_SA_JSON"]),
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"],
    )
    sh = gspread.authorize(creds).open_by_key(os.environ["GOOGLE_SHEET_ID"])
    tab_name = "s8e8_finance_raw"
    fieldnames = ['market', 'year', 'month', 'month_num', 'patients',
                  'ad_spend', 'collections', 'cost_per_patient',
                  'source_row', 'block_width', 'notes']
    try:
        ws = sh.worksheet(tab_name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=tab_name, rows=len(all_rows) + 10, cols=len(fieldnames))
    data = [fieldnames] + [[r.get(h, '') for h in fieldnames] for r in all_rows]
    ws.update(data, value_input_option="RAW")
    print(f"  ✓ wrote {len(all_rows)} rows to Sheet tab '{tab_name}'")


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="Path to the s8e8 Marketing ROI Report xlsx")
    ap.add_argument("--out-csv", help="Also/instead write a local CSV copy (gitignored, for inspection only)")
    ap.add_argument("--skip-sheet", action="store_true", help="Don't write to the Google Sheet (just --out-csv)")
    ap.add_argument("--log", default=str(DEFAULT_LOG))
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    all_rows = []
    full_log = {}
    for sheet_name, market_name in MARKET_MAP.items():
        log = {'sheet_name': sheet_name}
        rows, log = process_sheet(wb, sheet_name, market_name, log)
        all_rows.extend(rows)
        full_log[market_name] = log

    all_rows.sort(key=lambda r: (r['market'], r['year'], r['month_num']))

    if args.out_csv:
        with open(args.out_csv, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=['market', 'year', 'month', 'month_num', 'patients',
                                                    'ad_spend', 'collections', 'cost_per_patient',
                                                    'source_row', 'block_width', 'notes'])
            writer.writeheader()
            for r in all_rows:
                writer.writerow(r)
        print(f"Wrote {len(all_rows)} rows to {args.out_csv}")

    if not args.skip_sheet:
        write_to_sheet(all_rows)

    with open(args.log, 'w') as f:
        json.dump(full_log, f, indent=2, default=str)

    for market, log in full_log.items():
        print('---', market, '---')
        for k, v in log.items():
            if k in ('year_markers',):
                continue
            print(' ', k, ':', v)


if __name__ == '__main__':
    main()
